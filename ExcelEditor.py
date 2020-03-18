import openpyxl
from openpyxl.styles import Alignment
import datetime
import smtplib


class ExcelEditor:

    def __init__(self, excelWorkbook, sheet, min_row, max_col, emailUsername, emailPassword): 
        self.excelWorkbook = excelWorkbook
        self.wb = openpyxl.load_workbook(self.excelWorkbook)
        self.sheet = self.wb[sheet] 
        self.min_row = min_row  # the starting row where data will be written to in the excel page
        self.max_col = max_col  # the starting column where data will be written to in the excel page
        self.dict_of_links = self.GetExcelEntries()
        self.emailUsername = emailUsername
        self.emailPassword = emailPassword
        self.backupWorkbook = 'BACKUP_Workbook.xlsx'

    def GetExcelEntries(self):
        dict_of_links = {}
        currentRow = self.min_row  # the row of products starts at min_row
        for row in self.sheet.iter_rows(min_row=self.min_row, max_col=self.max_col, max_row=self.sheet.max_row, values_only=True):  
            excel_websiteLink = row[2]
            dict_of_links.update({excel_websiteLink : currentRow})  # add each item's link in the excel sheet and their respective row to a dictionary(Key,Value)
            currentRow = currentRow + 1
        return dict_of_links
        
    def AddToExcel(self, productName, excelPrice, webLink):
        """ Adds the item to the excel sheet if the item is new, otherwise update's the item's price
        Args:
        productName: the name of the product
        excelPrice: the price that is saved in the excel sheet and used as comparison for a deal
        webLink: the url link of the product's store page/website
        """
        currentDT = datetime.datetime.now()

        if webLink not in self.dict_of_links:
            print('*NEW ITEM* ADDING ' + productName + ' TO EXCEL SHEET!')
            self.sheet.cell(row=self.sheet.max_row+1, column=1, value=productName)
            self.sheet.cell(row=self.sheet.max_row, column=2, value=excelPrice).alignment = Alignment(horizontal='center') 
            self.sheet.cell(row=self.sheet.max_row, column=3, value=webLink)
            self.sheet.cell(row=self.sheet.max_row, column=4, value=currentDT.strftime("%b %d %Y")).alignment = Alignment(horizontal='center')
            self.dict_of_links.update({webLink : self.sheet.max_row}) 
        elif webLink in self.dict_of_links:
            print('*ITEM ALREADY EXISTS* UPDATED ' + productName + '!')
            self.UpdateExcel(webLink, excelPrice)
          
        self.wb.save(self.excelWorkbook) 
        self.wb.save(self.backupWorkbook)

    def UpdateExcel(self, webLink, newPrice):
        """ Updates the item's price and date inside the excel sheet
        Args:
        webLink: the url link of the product's store page/website
        newPrice: the new price that will replace the old price in the excel sheet
        """
        currentDT = datetime.datetime.now()
        row = self.dict_of_links.get(webLink)  # gets the row in the excel page where the product is already stored
        self.sheet.cell(row=row, column=2, value=newPrice) 
        self.sheet.cell(row=row, column=4, value=currentDT.strftime("%b %d %Y"))
        self.wb.save(self.excelWorkbook)
        self.wb.save(self.backupWorkbook)   

    def SendEmail(self, FROM, TO, SUBJECT, TEXT):
        emailUsername = self.emailUsername
        emailPassword = self.emailPassword
        smtpObj = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        smtpObj.ehlo()
        smtpObj.login(emailUsername, emailPassword)
        smtpObj.sendmail(FROM, TO, 'Subject: {}\n\n{}'.format(SUBJECT,TEXT))
        smtpObj.quit()        
