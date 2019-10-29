from bs4 import BeautifulSoup
import requests
import smtplib
import openpyxl
import datetime
import re
import time
import sys
import PriceCheckerGUI as PCG


sender_emailUsername = 'tom@example.com'  #  Email user will use as the sender for the notification email
sender_emailPassword = 'password123'
receiver_emailUsername = 'will@example.com'  # Email address where user will receieve the notification email
wb = openpyxl.load_workbook('Book1.xlsx')
sheet = wb['Sheet']


def AddToExcel(productName, currentPrice, webLink, webName, excelPrice, dealType):
    """ Adds the item to the excel sheet if the item is new or if it the item already exists, updates the price if there's a better deal
    Args:
      productName: the name of the product
      currentPrice: the price of the product that it is currently being sold for
      webLink: the url link of the product's store page
      webName: the name of the website
      excelPrice: the price that is saved in the excel sheet and used as comparison for a deal
      dealType: determines whether current price is a better deal than saved excel price
    """
    currentDT = datetime.datetime.now()
    items_in_excel_sheet = []
    for row in sheet.iter_rows(min_row=2, max_col=4, max_row=sheet.max_row, values_only=True):  # add each item in the excel sheet to an array
        items_in_excel_sheet.append(row)

    new_item = True  # item being added is a new item not in the excel sheet
    for item in items_in_excel_sheet:
        currentRow = items_in_excel_sheet.index(item) + 2  # the excel sheet starts at a different index than the array
        #excel_productName = item[0] 
        #excel_productPrice = item[1] 
        excel_websiteLink = item[2]
        #excel_websiteName = item[3]

        if webLink in excel_websiteLink and dealType==0:  # if current price > than saved excel price
            result = 'FOUND ' + productName + ' WORSE PRICE: (BEST_PRICE = ' + str(excelPrice) +') ' + '(CURRENT_PRICE = ' + str(currentPrice)+')'
            print(result)
            PCG.InsertText(result, color=1)
            new_item = False
        elif webLink in excel_websiteLink and dealType==1:  # if current price == to saved excel price
            result = 'FOUND ' + productName + ' WITH EQUAL PRICE: (BEST_PRICE = ' + str(excelPrice) +') ' + '(CURRENT_PRICE = ' + str(currentPrice)+')'
            print(result)
            PCG.InsertText(result, color=2)
            new_item = False
        elif webLink in excel_websiteLink and dealType==2:  # if current price is < than saved excel price (there's a sale/deal)
            result = 'FOUND ' + productName + ' WITH BETTER PRICE: (OLD_BEST_PRICE = ' + str(excelPrice) +') ' + '(CURRENT_PRICE = ' + str(currentPrice)+')'
            print(result)
            PCG.InsertText(result, color=3)
            sheet.cell(row=currentRow, column=1, value=productName) 
            sheet.cell(row=currentRow, column=2, value=currentPrice)  
            sheet.cell(row=currentRow, column=5, value=currentDT.strftime("%b %d %Y"))
            new_item = False
    if new_item == True:
        result = '*NEW ITEM* ADDING ' + productName + ' TO EXCEL SHEET!'
        print(result)
        PCG.InsertText(result, color=3)
        sheet.cell(row=sheet.max_row+1, column=1, value=productName) 
        sheet.cell(row=sheet.max_row, column=2, value=excelPrice) 
        sheet.cell(row=sheet.max_row, column=3, value=webLink)
        sheet.cell(row=sheet.max_row, column=4, value=webName)
        sheet.cell(row=sheet.max_row, column=5, value=currentDT.strftime("%b %d %Y"))

    PCG.output_text.see('end')
    wb.save('Book1.xlsx') 
    wb.save('Book1_Backup.xlsx')

def CheckPrices():
    """ Checks the prices of all the items in the excel sheet for a better deal
    """
    items_in_excel_sheet = []
    for row in sheet.iter_rows(min_row=2, max_col=4, max_row=sheet.max_row, values_only=True):
        items_in_excel_sheet.append(row)
    
    for item in items_in_excel_sheet:
        excel_productPrice = item[1]
        excel_websiteLink = item[2]
        excel_websiteName = item[3]

        if excel_websiteName == '93Brand':
            print('93Brand Found:  ', excel_websiteLink)
            Parse_93Brand(excel_websiteLink, excel_productPrice)
        elif excel_websiteName == 'Adidas':
            print('Adidas FOUND:  ', excel_websiteLink) 
            Parse_Adidas(excel_websiteLink, excel_productPrice)
        elif excel_websiteName == 'BananaRepublic':
            print('BananaRepublic FOUND:  ', excel_websiteLink)
            Parse_BananaRepublic(excel_websiteLink, excel_productPrice)
        elif excel_websiteName == 'FightersMarket':
            print('FightersMarket FOUND:  ', excel_websiteLink)
            Parse_FightersMarket(excel_websiteLink, excel_productPrice)
        elif excel_websiteName == 'Microcenter':
            print('Microcenter FOUND:  ', excel_websiteLink)
            Parse_Microcenter(excel_websiteLink, excel_productPrice)
        elif excel_websiteName == 'Nike':
            print('Nike FOUND:  ', excel_websiteLink)
        
    wb.save('Book1.xlsx')  
    wb.save('Book1_Backup.xlsx')  
    print('FINISHED CHECKING PRICES...')
    PCG.InsertText('FINISHED CHECKING PRICES...', color=0)

""" The following Parse functions vary slightly since every website is built differently.
Each Parse function uses requests to connect to the product's website and then the html is parsed using
Beautiful Soup for the product's name and price. The parse info is passed to the BestPrice function to 
determine whether it's a better deal than the price saved in the excel sheet. """
def Parse_93Brand(webLink, excelPrice):

    webName='93Brand'
    header_info = {'user-agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36'}
    r = requests.get(webLink, headers=header_info)

    if r.status_code == 200:
        try:
            soup = BeautifulSoup(r.text, 'lxml')
            time.sleep(2)
            productName_temp1 = soup.find('div', class_='product-title')
            productName_temp2 =  productName_temp1.find('h1', itemprop='name')
            productName = productName_temp2.text.strip()
            match1 = soup.find('div', class_='price')
            match2 = match1.find('span', class_='money')
            currentPrice = match2.text
            currentPrice = re.findall(r"[-+]?\d*\.\d+|\d+", currentPrice) #removes the $ sign, if the product has a sale price, it will take both prices and put it into an array where index 0 = sale price
            currentPrice = float(currentPrice[0])
            BestPrice(productName, currentPrice, webLink, webName, excelPrice)
        except AttributeError:
            print('ATTRIBUTE ERROR FOUND ON: ', webName)
    else:
        print('ERROR LOADING WEBSITE [' + webName + '] ' + 'WITH STATUS_CODE: ' + str(r.status_code))

def Parse_Adidas(webLink, excelPrice):

    webName='Adidas'
    header_info = {'user-agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36'}
    r = requests.get(webLink, headers=header_info)

    if r.status_code == 200:
        try:
            soup = BeautifulSoup(r.text, 'lxml')
            time.sleep(3)
            productName = soup.find('h1', class_='product_title')
            productName = productName.text
            productColor = soup.find('div', class_='color_text___mgoYV')
            productColor = productColor.text
            productName = productName + ' (Color: ' + productColor + ')'
            match1 = soup.find('div', class_='gl-price')
            currentPrice = match1.text
            currentPrice = re.findall(r"[-+]?\d*\.\d+|\d+", currentPrice) #removes the $ sign, if the product has a sale price, it will take both prices and put it into an array where index 0 = sale price
            currentPrice = float(currentPrice[0])
            BestPrice(productName, currentPrice, webLink, webName, excelPrice)
        except AttributeError:
            print('ATTRIBUTE ERROR FOUND ON: ', webName)
    else:
        print('ERROR LOADING WEBSITE [' + webName + '] ' + 'WITH STATUS_CODE: ' + str(r.status_code))

def Parse_BananaRepublic(webLink, excelPrice):

    webName='BananaRepublic'
    #header_info = {'user-agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36'}
    #r = requests.get(webLink, headers=header_info)
    r = requests.get(webLink)

    if r.status_code == 200:
        try:
            soup = BeautifulSoup(r.text, 'lxml')
            time.sleep(2)
            productName_temp =  soup.find('h1', class_='product-title__text')
            productName = productName_temp.text
            match1 = soup.find('h2', class_='product-price--pdp')
            currentPrice = match1.text
            currentPrice = re.findall(r"[-+]?\d*\.\d+|\d+", currentPrice) #removes the $ sign, if the product has a sale price, it will take both prices and put it into an array where index 0 = sale price
            currentPrice = float(currentPrice[0])
            BestPrice(productName, currentPrice, webLink, webName, excelPrice)
        except AttributeError:
            print('ATTRIBUTE ERROR FOUND ON: ', webName)
    else:
        print('ERROR LOADING WEBSITE [' + webName + '] ' + 'WITH STATUS_CODE: ' + str(r.status_code))
    
def Parse_FightersMarket(webLink, excelPrice):

    webName='FightersMarket'
    header_info = {'user-agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36'}
    r = requests.get(webLink, headers=header_info)

    if r.status_code == 200:
        try:
            soup = BeautifulSoup(r.text, 'lxml')
            time.sleep(2)
            productName_temp1 = soup.find('div', class_='product-title')
            productName_temp2 =  productName_temp1.find('h1', itemprop='name')
            productName = productName_temp2.text
            match1 = soup.find('div', class_='price_money')
            match2 = match1.find('span', class_='money')
            currentPrice = match2.text
            currentPrice = re.findall(r"[-+]?\d*\.\d+|\d+", currentPrice) #removes the $ sign, if the product has a sale price, it will take both prices and put it into an array where index 0 = sale price
            currentPrice = float(currentPrice[0])
            BestPrice(productName, currentPrice, webLink, webName, excelPrice)
        except AttributeError:
            print('ATTRIBUTE ERROR FOUND ON: ', webName)
    else:
        print('ERROR LOADING WEBSITE [' + webName + '] ' + 'WITH STATUS_CODE: ' + str(r.status_code))

def Parse_Microcenter(webLink, excelPrice):

    webName='Microcenter'
    header_info = {'user-agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36'}
    r = requests.get(webLink,headers=header_info)

    if r.status_code == 200:
        try:
            soup = BeautifulSoup(r.text, 'lxml')
            time.sleep(2)
            match1 = soup.find('div', id='details')
            match2 = [item['data-brand'] for item in match1.find_all() if "data-brand" in item.attrs] #item brand
            match3 = [item['data-name'] for item in match1.find_all() if "data-name" in item.attrs] #item name
            productName = match2[0] + ' ' + match3[0]
            match1 = soup.find('span', id='pricing')
            currentPrice = match1.text
            currentPrice = re.findall(r"[-+]?\d*\.\d+|\d+", currentPrice) #removes the USD after the price (just want the value)
            currentPrice = float(currentPrice[0])
            BestPrice(productName, currentPrice, webLink, webName, excelPrice)
        except AttributeError:
            print('ATTRIBUTE ERROR FOUND ON: ', webName)
    else:
        print('ERROR LOADING WEBSITE [' + webName + '] ' + 'WITH STATUS_CODE: ' + str(r.status_code))

def BestPrice(productName, currentPrice, webLink, webName, excelPrice):
    if currentPrice > excelPrice:
        AddToExcel(productName, currentPrice, webLink, webName, excelPrice, dealType=0)
    elif currentPrice == excelPrice:
        AddToExcel(productName, currentPrice, webLink, webName, excelPrice, dealType=1)
    elif currentPrice < excelPrice:
        AddToExcel(productName, currentPrice, webLink, webName, excelPrice, dealType=2)
        SendEmail(sender_emailUsername, receiver_emailUsername, productName + ': ' + str(currentPrice), productName + ' is on sale for: $' + str(currentPrice) +' on the following website: ' + webLink)
        PCG.InsertText('EMAIL SENT! THIS ITEM IS ON SALE!', color=4)
        print('EMAIL SENT!')  

def SendEmail(FROM, TO, SUBJECT, TEXT):
    smtpObj = smtplib.SMTP_SSL('smtp.gmail.com', 465)  # Edit this if you will be using a different email service provider
    smtpObj.ehlo()
    smtpObj.login(sender_emailUsername, sender_emailPassword)
    smtpObj.sendmail(FROM, TO, 'Subject: {}\n\n{}'.format(SUBJECT,TEXT))
    smtpObj.quit()              

def Menu():
    while True:
        try:
            print('\nWELCOME! Price Checker has started!')
            choice = input('1. Add a new item. \n2. Check Prices of Products \n3. Exit\nEnter your choice [1-3]: ')
            choice = int(choice)
            if choice == 1:
                    print('Choose the Website where the item is sold at:')
                    print(' 1a) Adidas\n 1b) BananaRepublic\n 1c) Microcenter')
                    choice2 = input()
                    if choice2 =='1a':
                        excel_websiteLink = input('Enter the product\'s link:  ')
                        excel_productPrice = input('Enter the product\'s price [Enter digits only Without $, USD, etc]:  ')
                        excel_productPrice = float(excel_productPrice)
                        Parse_Adidas(excel_websiteLink, excel_productPrice)
                    elif choice2 == '1b':
                        excel_websiteLink = input('Enter the product\'s link:  ')
                        excel_productPrice = input('Enter the product\'s price [Enter digits only Without $, USD, etc]:  ')
                        excel_productPrice = float(excel_productPrice)
                        Parse_BananaRepublic(excel_websiteLink, excel_productPrice)
                    elif choice2 == '1c':
                        excel_websiteLink = input('Enter the product\'s link:  ')
                        excel_productPrice = input('Enter the product\'s price [Enter digits only Without $, USD, etc]:  ')
                        excel_productPrice = float(excel_productPrice)
                        Parse_Microcenter(excel_websiteLink, excel_productPrice)
                    else:
                        print('Invalid number. Try again...')
            elif choice == 2:
                    CheckPrices()
            elif choice == 3:
                    print ('GOODBYE!')
                    break
            else:
                print('Invalid number. Try again...')
        except ValueError:
            print("Oops!  That was an invalid entry.  Try again...")


if __name__ == "__main__":
    """ Running the python file without any arguments will start the application 
    however running the program with an additional argument (doesn't matter what it is) will 
    start the program's CheckPrices() function. This is so that I can use the Windows Task Scheduler to 
    check for prices at a given time and so that the user doesn't have to wait for CheckPrices() to conclude
    at the beginning of each startup.
    """
    if len(sys.argv) == 1:
        PCG.root.mainloop()
    elif len(sys.argv) == 2:
        CheckPrices()
        PCG.root.mainloop()
    else:
        pass